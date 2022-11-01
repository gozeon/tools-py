from datetime import datetime
import os
from turtle import title
from openpyxl import load_workbook
from copy import copy
import requests


api_url = os.getenv("api_url")
project_id = os.getenv("project_id")
author = os.getenv("author")

# 获取项目
response = requests.get("{}/api/project/{}".format(api_url, project_id))
project_res = response.json()
project_info = project_res['data']

# 获取记录
response = requests.get("{}/api/record/{}".format(api_url, project_id))
record_res = response.json()
record_info = record_res['data']

title = project_info['title'].replace("自测", "")

insert_line = len(record_info) + 1
min_row = 8
end_row = insert_line + min_row
wb = load_workbook("templates/temp.xlsx")

ws_index = wb["首页"]

ws_index['D4'] = title
ws_index['D6'] = author

ws_test = wb["自测"]

# 插入行
merged_cells_range = ws_test.merged_cells.ranges
for merged_cell in merged_cells_range:
    if merged_cell.min_row > min_row:
        merged_cell.shift(0, insert_line)

ws_test.insert_rows(min_row + 1, insert_line)

ws_test.unmerge_cells(range_string="B6:B8")
ws_test.merge_cells(range_string="B6:B" + str(end_row))


# 格式化样式
for row_i in range(6, end_row):
    # 设置高度
    ws_test.row_dimensions[row_i +
                           1].height = ws_test.row_dimensions[row_i].height
    for col_j in list(map(chr, range(66, 74))):
        # C - I
        old_cell = ws_test["{column}{row}".format(row=row_i, column=col_j)]
        next_row_cell = ws_test["{column}{row}".format(
            row=row_i + 1, column=col_j)]
        if old_cell.has_style:
            # next_row_cell.font = copy(old_cell.font)
            # next_row_cell.border = copy(old_cell.border)
            # next_row_cell.fill = copy(old_cell.fill)
            # next_row_cell.number_format = copy(old_cell.number_format)
            # next_row_cell.protection = copy(old_cell.protection)
            # next_row_cell.alignment = copy(old_cell.alignment)
            next_row_cell._style = copy(old_cell._style)

# 插入内容
for index, record in enumerate(record_info):
    curr_row = str(6 + index)
    ws_test["C" + curr_row] = record['title']
    ws_test["E" + curr_row] = datetime.fromisoformat(record['CreatedAt']).strftime("%Y/%m/%d")
    ws_test["G" + curr_row] =author

    if record['status'] == 2:
        ws_test["F" + curr_row] = "PASS"
    else:
        ws_test["F" + curr_row] = "FAIL"
        ws_test["H" + curr_row] = "FAIL"
        if record['status'] > 3:
            ws_test["I" + curr_row] = "该用例已被取代或弃用"
        else:
            ws_test["I" + curr_row] = "无法准备相关用例"


file_name = "{}.xlsx".format(project_info['title'])

if os.path.exists(file_name):
    os.remove(file_name)

wb.save(file_name)
wb.close()
